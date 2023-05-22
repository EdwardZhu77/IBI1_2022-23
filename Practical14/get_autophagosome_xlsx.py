# Import modules
import xml.dom.minidom as dom  
import openpyxl  
import os
os.chdir("E:\study files (python)\IBI\IBI Practical")# To work on my laptop'''

# Define a new function to work on it
def get_term_info(term): 
    '''Extract GO ID, name, definition from a term'''
    go_id = term.getElementsByTagName('id')[0].childNodes[0].data  # Get the GO ID and term name
    name = term.getElementsByTagName('name')[0].childNodes[0].data
    defstr = term.getElementsByTagName('defstr')[0].childNodes[0].data # Get the definition
    return go_id, name, defstr

# Define another new function to work on it
def find_child_nodes(term, all_terms):  
    '''Find all child nodes of a term'''
    child_nodes = []  # Define an empty list to store child nodes
    term_id = term.getElementsByTagName('id')[0].childNodes[0].data # Get the ID of the current term
    for t in all_terms:  # For each term t in the list of all terms
        is_a = t.getElementsByTagName('is_a')  # Get all 'is_a' tags for the current term t
        for isa in is_a:  # For each 'is_a' tag
            if isa.childNodes[0].data == term_id: # Check if the ID matches the term we are looking for
                child_nodes.append(t) # If so, we found a child node. Add it to the list
                child_nodes.extend(find_child_nodes(t, all_terms)) # Recursively call the function to find the child nodes of the child node we just found
    return child_nodes  

# Load XML 
xml_doc = dom.parse('go_obo.xml')  
terms = xml_doc.getElementsByTagName('term')  

# Create spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'GO ID'  
sheet['B1'] = 'Name'
sheet['C1'] = 'Definition' 
sheet['D1'] = 'Number of child nodes'

row = 2
for term in terms:
    # Get info from term
    go_id, name, defstr = get_term_info(term)
    
    # Check if relevant
    if 'autophagosome' in defstr:     
        
        # Find child nodes
        child_nodes = find_child_nodes(term, terms)  
        num_child_nodes = len(child_nodes)  
        
        # Add to spreadsheet
        sheet.cell(row=row, column=1).value = go_id  
        sheet.cell(row=row, column=2).value = name
        sheet.cell(row=row, column=3).value = defstr 
        sheet.cell(row=row, column=4).value = num_child_nodes  
        
        row += 1
        
wb.save('autophagosome.xlsx')